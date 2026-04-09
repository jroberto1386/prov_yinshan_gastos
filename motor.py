# ─────────────────────────────────────────────
#  motor.py  —  Provisión de Gastos / Proveedores
#  Fuente: Excel de facturas recibidas + catálogo
# ─────────────────────────────────────────────

import io
import re
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill

from config import (
    CTA_IVA_16, CTA_IVA_8, CTA_IEPS,
    CTA_RET_IVA, CTA_RET_ISR_GENERAL, CTA_RET_ISR_ARRENDAM,
    CTA_GASTO_DEFAULT, PALABRAS_ARRENDAMIENTO,
    TIPO_POL, ID_DIARIO,
)

# ── 22 filas de encabezado CONTPAq (mismo schema que provisión ingresos) ──
HEADERS = [
    ["Egreso(EG)", "IdDocumentoDe", "TipoDocumento", "Folio", "Fecha", "FechaAplicacion",
     "CodigoPersona", "BeneficiarioPagador", "IdCuentaCheques", "CodigoMoneda",
     "Total", "Referencia", "Origen", "BancoDestino", "CuentaDestino",
     "OtroMetodoDePago", "Guid", None, None, "TipoCambio",
     "UUIDRep", "NodoPago", "CodigoMonedaTipoCambio", "NumAsoc"],
    ["deposito.1(DE)", "IdDocumentoDe", "TipoDocumento", "Folio", "Fecha", "Ejercicio",
     "Periodo", "FechaAplicacion", "EjercicioAp", "PeriodoAp",
     "IdCuentaCheques", "NatBancaria", "Naturaleza", "Total", "Referencia",
     "Concepto", "EsConciliado", "IdMovEdoCta", "EjercicioPol", "PeriodoPol",
     "TipoPol", "NumPol", "FormaDeposito", "IdPoliza", "Origen",
     "IdDocumento", "PolizaAgrupada", "UsuarioCrea", "UsuarioModifica", "tieneCFD", "Guid"],
    ["ingreso.1(IN)", "IdDocumentoDe", "TipoDocumento", "Folio", "Fecha", "FechaAplicacion",
     "CodigoPersona", "BeneficiarioPagador", "IdCuentaCheques", "CodigoMoneda",
     "Total", "Referencia", "Origen", "BancoOrigen", "CuentaOrigen",
     "OtroMetodoDePago", "Guid", None, None, "TipoCambio",
     "NumeroCheque", "UUIDRep", "NodoPago", "CodigoMonedaTipoCambio", "NumAsoc"],
    ["Datos para CONTPAQi Factura Electrónica®(FE)", "RutaAnexo", "ArchivoAnexo"],
    ["Movimiento de póliza(M1)", "IdCuenta", "Referencia", "TipoMovto", "Importe",
     "IdDiario", "ImporteME", "Concepto", "IdSegNeg", "Guid", "FechaAplicacion"],
    ["Devolución de IVA (IETU)(W)", "IETUDeducible", "IETUModificado"],
    ["Devolución de IVA(V)", "IdProveedor", "ImpTotal", "PorIVA", "ImpBase",
     "ImpIVA", "CausaIVA", "ExentoIVA", "Serie", "Folio", "Referencia",
     "OtrosImptos", "ImpSinRet", "IVARetenido", "ISRRetenido", "GranTotal",
     "EjercicioAsignado", "PeriodoAsignado", "IdCuenta", "IVAPagNoAcred",
     "UUID", None, "IEPS"],
    ["Asociación de nodo de pago(AP)", "UUIDRep", "NumNodoPago", "GuidReferencia", "AplicationType"],
    ["Periodo de causación de IVA(R)", "EjercicioAsignado", "PeriodoAsignado"],
    ["Póliza(P)", "Fecha", "TipoPol", "Folio", "Clase", "IdDiario",
     "Concepto", "SistOrig", "Impresa", "Ajuste", "Guid"],
    ["Asociación movimiento(AM)", "UUID"],
    ["Comprobantes(MC)", "IdCuentaFlujoEfectivo", "IdSegmentoNegCtaFlujo", "Fecha",
     "Serie", "Folio", "UUID", "ClaveRastreo", "Referencia", "IdProveedor",
     "CodigoConceptoIETU", "ImpNeto", "ImpNetoME", "IdCuentaNeto",
     "IdSegmentoNegNeto", "PorIVA", "ImporteIVA", "ImporteIVAME",
     "IVATasaExcenta", "IdCuentaIVA", "IdSegmentoNegIVA", "NombreImpuesto",
     "ImpImpuesto", "ImpImpuestoME", "IdCuentaImpuesto", "IdSegmentoNegImp",
     "ImpOtrosGastos", "ImpOtrosGastosME", "IdCuentaOtrosGastos",
     "IdSegmentoNegOtrosGastos", "IVARetenido", "IVARetenidoME",
     "IdCuentaRetIVA", "IdSegmentoNegRetIVA", "ISRRetenido", "ISRRetenidoME",
     "IdCuentaRetISR", "IdSegmentoNegRetISR", "NombreOtrasRetenciones",
     "ImpOtrasRetenciones", "ImpOtrasRetencionesME", "IdCuentaOtrasRetenciones",
     "IdSegmentoNegOtrasRet", "BaseIVADIOT", "BaseIETU", "IVANoAcreditable",
     "ImpTotalErogacion", "IVAAcreditable", "ImpExtra1", "ImpExtra2",
     "IdCategoria", "IdSubCategoria", "TipoCambio", "IdDocGastos",
     "EsCapturaCompleta", "FolioStr"],
    ["Movimiento de póliza(M)", "IdCuenta", "Referencia", "TipoMovto", "Importe",
     "IdDiario", "ImporteME", "Concepto", "IdSegNeg"],
    ["Dispersiones de pago(DP)", "UUID", "UUIDRep", "GuidRef", "NumNodoPago",
     "FechaPago", "TotalPago", "TipoCambio", "TotalPagoComprobante"],
    ["Devolución de IVA (IETU)(W2)", "IETUDeducible", "IETUAcreditable",
     "IETUModificado", "IdConceptoIETU"],
    ["Movimientos de impuestos(I)", "IdPersona", "EjercicioAsignado",
     "PeriodoAsignado", "IdCuenta", "AplicaImpuesto", "Serie", "Folio",
     "Referencia", "UUID", "Origen", "Computable", "TipoMovimiento",
     "TipoFactor", "Impuesto", "ObjetoImpuesto", "NombreImpLocal",
     "TasaOCuota", "ImpBase", "ImpImpuesto", "ImpTotal", "Desglosado",
     "IVANoAcred", "AcumulaIETU", "IdConceptoIETU", "IETUDeducible",
     "IETUModificado", "IETUAcreditable", "GuidMov", "GuidMovPadre",
     "Migrado", "ConceptoIVA", "SubconceptoIVA", "ClasificadorIVA",
     "ProporcionDIOT", "DeducibleDIOT"],
    ["Asociación documento(AD)", "UUID"],
    ["Cheque(CH)", "IdDocumentoDe", "TipoDocumento", "Folio", "Fecha",
     "FechaAplicacion", "CodigoPersona", "BeneficiarioPagador",
     "IdCuentaCheques", "CodigoMoneda", "Total", "Referencia", "Origen",
     "CuentaDestino", "BancoDestino", "Guid", None, "OtroMetodoDePago",
     "TipoCambio", "UUIDRep", "NodoPago", "CodigoMonedaTipoCambio", "NumAsoc"],
    ["IngresosNoDepositados.1(DI)", "IdDocumentoDe", "TipoDocumento", "Folio",
     "Fecha", "Ejercicio", "Periodo", "FechaAplicacion", "EjercicioAp",
     "PeriodoAp", "CodigoPersona", "BeneficiarioPagador", "NatBancaria",
     "Naturaleza", "CodigoMoneda", "CodigoMonedaTipoCambio", "TipoCambio",
     "Total", "Referencia", "Concepto", "EsAsociado", "UsuAutorizaPresupuesto",
     "PosibilidadPago", "EsProyectado", "Origen", "IdChequeOrigen",
     "TipoCambioDeposito", "IdDocumento", "EsAnticipo", "EsTraspasado",
     "UsuarioCrea", "UsuarioModifica", "tieneCFD", "Guid",
     "CuentaOrigen", "BancoOrigen", "OtroMetodoDePago", "NumeroCheque", "NumAsoc"],
    ["Causación de IVA (Concepto de IETU)(E)", "IdConceptoIETU"],
    ["Causación de IVA (IETU)(D)", "IVATasa15NoAcred", "IVATasa10NoAcred",
     "IETU", "Modificado", "Origen", "TotTasa16", "BaseTasa16", "IVATasa16",
     "IVATasa16NoAcred", "TotTasa11", "BaseTasa11", "IVATasa11",
     "IVATasa11NoAcred", "TotTasa8", "BaseTasa8", "IVATasa8", "IVATasa8NoAcred"],
    ["Causación de IVA(C)", "Tipo", "TotTasa15", "BaseTasa15", "IVATasa15",
     "TotTasa10", "BaseTasa10", "IVATasa10", "TotTasa0", "BaseTasa0",
     "TotTasaExento", "BaseTasaExento", "TotOtraTasa", "BaseOtraTasa",
     "IVAOtraTasa", "ISRRetenido", "TotOtros", "IVARetenido",
     "Captado", "NoCausar", "IEPS"],
]


# ── Helpers ───────────────────────────────────

def _f(val, default=0.0):
    """Convierte a float seguro."""
    try:
        v = float(val)
        return v if not pd.isna(v) else default
    except (TypeError, ValueError):
        return default


def _s(val, default=""):
    """Convierte a str seguro."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return default
    return str(val).strip()


def es_arrendamiento(concepto_texto):
    """Detecta si el concepto describe un arrendamiento."""
    texto = concepto_texto.lower()
    return any(p in texto for p in PALABRAS_ARRENDAMIENTO)


# ── Carga del catálogo de proveedores ─────────

def cargar_catalogo(catalogo_bytes):
    """
    Lee el Excel/XLS del catálogo y devuelve:
      - dict RFC → {cta_proveedor, cta_gasto, nombre, codigo}
      - ultimo_codigo (int)
    Soporta .xls (legacy) y .xlsx automáticamente.
    """
    # Intentar openpyxl (xlsx) primero; si falla, convertir con LibreOffice
    try:
        df = pd.read_excel(io.BytesIO(catalogo_bytes), header=None, engine="openpyxl")
    except Exception:
        # Es un .xls legacy — convertir a xlsx en memoria con LibreOffice
        import tempfile, subprocess, os
        with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as tmp:
            tmp.write(catalogo_bytes)
            tmp_path = tmp.name
        out_dir = tempfile.mkdtemp()
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "xlsx", "--outdir", out_dir, tmp_path],
            capture_output=True, timeout=60
        )
        os.unlink(tmp_path)
        xlsx_path = os.path.join(out_dir, os.path.basename(tmp_path).replace(".xls", ".xlsx"))
        with open(xlsx_path, "rb") as f:
            xlsx_bytes = f.read()
        os.unlink(xlsx_path)
        df = pd.read_excel(io.BytesIO(xlsx_bytes), header=None, engine="openpyxl")

    # Filtrar filas P1 (proveedores)
    provs = df[df[0] == "P1"].copy()

    catalogo = {}
    for _, r in provs.iterrows():
        rfc = _s(r[3])
        if not rfc:
            continue
        try:
            cta_prov  = int(float(r[6]))
        except (ValueError, TypeError):
            cta_prov  = None
        try:
            cta_gasto = int(float(r[26])) if not pd.isna(r[26]) else CTA_GASTO_DEFAULT
        except (ValueError, TypeError):
            cta_gasto = CTA_GASTO_DEFAULT

        try:
            codigo = int(float(r[1]))
        except (ValueError, TypeError):
            codigo = 0

        catalogo[rfc] = {
            "nombre":       _s(r[2]),
            "codigo":       codigo,
            "cta_proveedor": cta_prov,
            "cta_gasto":    cta_gasto,
        }

    # Último código de proveedor
    codigos = [v["codigo"] for v in catalogo.values() if v["codigo"] > 0]
    ultimo_codigo = max(codigos) if codigos else 0

    # Última cuenta secuencial real:
    # Tomamos la cuenta del proveedor con el CÓDIGO más alto,
    # filtrando cuentas en rango 201010001-201016000.
    # Excluye especiales: 201019999, 201022457, 205020001, etc.
    provs_num = provs.copy()
    provs_num["_codigo"] = pd.to_numeric(provs_num[1], errors="coerce")
    provs_num["_cuenta"] = pd.to_numeric(provs_num[6], errors="coerce")
    provs_seq = provs_num.dropna(subset=["_codigo", "_cuenta"])
    provs_seq = provs_seq[
        (provs_seq["_cuenta"] >= 201010001) &
        (provs_seq["_cuenta"] <= 201016000)
    ].sort_values("_codigo", ascending=False)
    ultima_cuenta = int(provs_seq.iloc[0]["_cuenta"]) if len(provs_seq) else 201015882

    return catalogo, ultimo_codigo, ultima_cuenta


# ── Procesamiento del Excel de facturas ────────

def procesar_excel(excel_bytes, catalogo_bytes, num_poliza_inicio, callback=None):
    """
    Lee el Excel de facturas recibidas, cruza con catálogo y devuelve:
      - facturas: lista de dicts con todos los campos para la póliza
      - nuevos_proveedores: lista de dicts para el layout de altas
      - stats: dict con conteos
    """
    catalogo, ultimo_codigo, ultima_cuenta = cargar_catalogo(catalogo_bytes)

    df = pd.read_excel(io.BytesIO(excel_bytes), header=None)
    data = df.iloc[1:].copy()  # saltar fila de headers
    total = len(data)

    # Primer pase: detectar todos los RFC nuevos y pre-asignarles cuenta consecutiva
    # Esto permite usarlas en las pólizas sin intervención manual posterior
    nuevos_provs_dict = {}   # RFC → {nombre, rfc, cta_proveedor}
    cuenta_nueva_idx  = [0]  # contador mutable para closures

    def obtener_cuenta_nueva(rfc, nombre):
        if rfc not in nuevos_provs_dict:
            nueva_cta = ultima_cuenta + len(nuevos_provs_dict) + 1
            nuevos_provs_dict[rfc] = {
                "nombre":        nombre,
                "rfc":           rfc,
                "cta_proveedor": nueva_cta,
            }
        return nuevos_provs_dict[rfc]["cta_proveedor"]

    facturas = []

    for i, (_, row) in enumerate(data.iterrows()):
        if callback:
            callback(i + 1, total)

        tipo = _s(row[3])
        # Solo Facturas y NotasCredito
        if tipo not in ("Factura", "NotaCredito"):
            continue

        es_nota = (tipo == "NotaCredito")
        signo   = -1 if es_nota else 1

        rfc_emisor = _s(row[12])
        nombre     = _s(row[13])
        serie      = _s(row[8])
        folio      = _s(row[9])
        uuid       = _s(row[10]).upper()
        concepto_txt = _s(row[40])

        # Fecha
        fecha_raw = row[4]
        try:
            if isinstance(fecha_raw, datetime):
                fecha = fecha_raw
            else:
                fecha = datetime.strptime(str(fecha_raw)[:10], "%Y-%m-%d")
        except (ValueError, TypeError):
            fecha = None

        # Importes
        subtotal    = _f(row[20])
        descuento   = _f(row[21])           # col V — Descuento
        ieps        = _f(row[22])           # Total IEPS (va a cuenta de gasto, no IVA)
        iva16_col   = _f(row[23])           # IVA 16% declarado en col X
        ret_iva     = _f(row[24])
        ret_isr     = _f(row[25])
        ish         = _f(row[26])           # ISH — Impuesto Sobre Hospedaje (col AA)
        total_cfdi  = _f(row[27])
        tot_traslad = _f(row[29])           # TotalTrasladados col AD (más confiable)
        iva8        = _f(row[56])

        # IVA real: cuando col23+col56 diverge de TotalTrasladados, usar col29
        iva_declarado = round(iva16_col + iva8, 2)
        if abs(iva_declarado - tot_traslad) > 0.05 and tot_traslad > 0:
            iva16 = round(tot_traslad - iva8, 2)
        else:
            iva16 = iva16_col

        # Gasto neto = SubTotal - Descuento
        # Nota: el IEPS (col 22) ya viene embebido dentro del SubTotal en el CFDI SAT,
        # por lo que NO se vuelve a sumar. ISH (col 26) sí es adicional.
        gasto_neto = round(subtotal - descuento + ish, 2)

        # Neto proveedor = Total del CFDI (el SAT ya descuenta las retenciones del total)
        neto_prov = round(total_cfdi, 2)

        # Referencia: F- + Serie + Folio
        # Si folio vacío → últimos 5 caracteres del UUID (sin guiones)
        folio_limpio = folio if folio and folio.lower() not in ("", "nan") else ""
        serie_limpia = serie if serie and serie.lower() not in ("", "nan") else ""
        if folio_limpio:
            ref = f"F-{serie_limpia}{folio_limpio}".strip()
        else:
            uuid_sin_guiones = uuid.replace("-", "")
            ref = f"F-{uuid_sin_guiones[-5:]}"

        # Buscar en catálogo
        prov = catalogo.get(rfc_emisor)
        if prov:
            cta_prov        = prov["cta_proveedor"]
            cta_gasto       = prov["cta_gasto"]
            nombre_catalogo = prov["nombre"] or nombre
        else:
            # Proveedor nuevo — asignar cuenta consecutiva pre-calculada
            cta_prov        = obtener_cuenta_nueva(rfc_emisor, nombre)
            cta_gasto       = CTA_GASTO_DEFAULT
            nombre_catalogo = nombre

        # Detectar arrendamiento
        arr = es_arrendamiento(concepto_txt)
        cta_ret_isr = CTA_RET_ISR_ARRENDAM if arr else CTA_RET_ISR_GENERAL

        # Concepto de la póliza P
        concepto_p = f"Provision {nombre_catalogo} {ref}"

        facturas.append({
            "fecha":           fecha,
            "rfc":             rfc_emisor,
            "nombre":          nombre_catalogo,
            "serie":           serie,
            "folio":           folio,
            "ref":             ref,
            "uuid":            uuid,
            "concepto_p":      concepto_p,
            "gasto_neto":      round(signo * gasto_neto, 2),   # Sub - Desc + IEPS + ISH
            "iva16":           round(signo * iva16, 2),
            "iva8":            round(signo * iva8, 2),
            "ret_iva":         round(signo * ret_iva, 2),
            "ret_isr":         round(signo * ret_isr, 2),
            "neto_prov":       round(signo * neto_prov, 2),    # Total CFDI (ya neto)
            "cta_gasto":       cta_gasto,
            "cta_prov":        cta_prov,
            "cta_ret_isr":     cta_ret_isr,
            "es_nota":         es_nota,
            "es_arrendam":     arr,
            "proveedor_nuevo": prov is None,
        })

    stats = {
        "procesadas":        len(facturas),
        "facturas":          sum(1 for f in facturas if not f["es_nota"]),
        "notas_credito":     sum(1 for f in facturas if f["es_nota"]),
        "con_iva8":          sum(1 for f in facturas if abs(f["iva8"]) > 0),
        "con_ret_iva":       sum(1 for f in facturas if abs(f["ret_iva"]) > 0),
        "con_ret_isr":       sum(1 for f in facturas if abs(f["ret_isr"]) > 0),
        "arrendamiento":     sum(1 for f in facturas if f["es_arrendam"]),
        "proveedores_nuevos": len(nuevos_provs_dict),
        "polizas_desde":     num_poliza_inicio,
        "polizas_hasta":     num_poliza_inicio + len(facturas) - 1,
    }

    return facturas, list(nuevos_provs_dict.values()), stats, ultimo_codigo


# ── Generación del Excel de Pólizas CONTPAq ───

def generar_polizas(facturas, num_poliza_inicio):
    """Genera el Excel de pólizas CONTPAq."""
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Pólizas Gastos"

    # Encabezados
    hfill = PatternFill("solid", fgColor="D9D9D9")
    hfont = Font(bold=True, size=9)
    for ri, hrow in enumerate(HEADERS, 1):
        for ci, v in enumerate(hrow, 1):
            c = ws.cell(ri, ci, v)
            c.font = hfont
            c.fill = hfill

    row_idx = 23
    num_pol = num_poliza_inicio

    for fact in facturas:
        fecha     = fact["fecha"]
        ref       = fact["ref"]
        nombre    = fact["nombre"]
        concepto  = fact["concepto_p"]
        cta_gasto = fact["cta_gasto"]
        cta_prov  = fact["cta_prov"]

        gasto_neto = fact["gasto_neto"]    # Sub - Desc + IEPS + ISH
        iva16      = fact["iva16"]
        iva8       = fact["iva8"]
        ret_iva    = fact["ret_iva"]
        ret_isr    = fact["ret_isr"]
        neto_prov  = fact["neto_prov"]     # Total CFDI (ya neto de retenciones)

        # ── Fila P ──
        p_vals = ["P", fecha, int(TIPO_POL), int(num_pol), 1, int(ID_DIARIO),
                  concepto, 12, 0, 0]
        for ci, v in enumerate(p_vals, 1):
            cell = ws.cell(row_idx, ci, v)
            if ci == 2 and isinstance(v, datetime):
                cell.number_format = "yyyymmdd"
        row_idx += 1

        # ── M1 Gasto (Debe) — SubTotal neto de descuento + IEPS + ISH ──
        _m1(ws, row_idx, cta_gasto, ref, 0, gasto_neto, nombre)
        row_idx += 1

        # ── M1 IVA 16% (Debe) ──
        if abs(iva16) > 0:
            _m1(ws, row_idx, CTA_IVA_16, ref, 0, iva16, nombre)
            row_idx += 1

        # ── M1 IVA 8% (Debe) ──
        if abs(iva8) > 0:
            _m1(ws, row_idx, CTA_IVA_8, ref, 0, iva8, nombre)
            row_idx += 1

        # ── M1 Retención IVA (Haber) ──
        if abs(ret_iva) > 0:
            _m1(ws, row_idx, CTA_RET_IVA, ref, 1, abs(ret_iva), nombre)
            row_idx += 1

        # ── M1 Retención ISR (Haber) ──
        if abs(ret_isr) > 0:
            _m1(ws, row_idx, fact["cta_ret_isr"], ref, 1, abs(ret_isr), nombre)
            row_idx += 1

        # ── M1 Proveedor (Haber) — neto ──
        # cta_prov siempre tiene valor: catálogo existente o cuenta nueva pre-asignada
        _m1(ws, row_idx, cta_prov, ref, 1, abs(neto_prov), nombre)
        row_idx += 1

        # ── AD — UUID ──
        if fact["uuid"]:
            ws.cell(row_idx, 1, "AD")
            ws.cell(row_idx, 2, fact["uuid"])
            row_idx += 1

        num_pol += 1

    # Ajuste ancho columnas
    for col in ws.columns:
        max_w = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_w + 2, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _m1(ws, row, cta, ref, tipo, importe, nombre):
    ws.cell(row, 1, "M1")
    ws.cell(row, 2, int(cta))
    ws.cell(row, 3, ref)
    ws.cell(row, 4, int(tipo))
    ws.cell(row, 5, importe)
    ws.cell(row, 6, int(ID_DIARIO))
    ws.cell(row, 7, 0)
    ws.cell(row, 8, nombre)


# ── Generación del Excel de Alta de Nuevos Proveedores ──

def generar_altas(nuevos_provs, ultimo_codigo, catalogo_bytes):
    """
    Genera un Excel en el mismo formato del catálogo de CONTPAq:
    - 6 filas de encabezado exactas del catálogo original
    - Nuevos proveedores desde la fila 7, con código y cuenta consecutivos
    - Cuenta de proveedor consecutiva a partir de la última del catálogo
    """
    # Leer el catálogo original para extraer las 6 filas de encabezado
    try:
        df_cat = pd.read_excel(io.BytesIO(catalogo_bytes), header=None, engine="openpyxl")
    except Exception:
        import tempfile, subprocess, os
        with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as tmp:
            tmp.write(catalogo_bytes)
            tmp_path = tmp.name
        out_dir = tempfile.mkdtemp()
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "xlsx", "--outdir", out_dir, tmp_path],
            capture_output=True, timeout=60
        )
        os.unlink(tmp_path)
        xlsx_path = os.path.join(out_dir, os.path.basename(tmp_path).replace(".xls", ".xlsx"))
        with open(xlsx_path, "rb") as f:
            xlsx_bytes = f.read()
        os.unlink(xlsx_path)
        df_cat = pd.read_excel(io.BytesIO(xlsx_bytes), header=None, engine="openpyxl")

    # Obtener la última cuenta del bloque secuencial real 201010001-201015999
    provs_cat = df_cat[df_cat[0] == "P1"]
    cuentas_prov = pd.to_numeric(provs_cat[6], errors="coerce").dropna()
    cuentas_seq  = cuentas_prov[(cuentas_prov >= 201010001) & (cuentas_prov <= 201015999)]
    ultima_cuenta = int(cuentas_seq.max()) if len(cuentas_seq) else 201015882

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nuevos Proveedores"

    hfill  = PatternFill("solid", fgColor="1F4E79")
    hfont  = Font(bold=True, color="FFFFFF", size=9)
    dfill  = PatternFill("solid", fgColor="D6E4F0")   # azul claro para encabezados de schema
    wfill  = PatternFill("solid", fgColor="FFF2CC")   # amarillo para filas de datos nuevos
    wfont  = Font(size=9)
    nfont  = Font(bold=True, color="C00000", size=9)  # rojo para nota de verificación

    # ── Filas 1-6: encabezados exactos del catálogo original ──
    for row_i in range(6):
        src_row = df_cat.iloc[row_i].tolist()
        for col_i, val in enumerate(src_row, 1):
            v = None if (isinstance(val, float) and pd.isna(val)) else val
            c = ws.cell(row_i + 1, col_i, v)
            c.font  = Font(bold=True, size=9)
            c.fill  = dfill

    # ── Filas 7+: nuevos proveedores ──
    # Columnas según schema P1 del catálogo (posiciones exactas):
    # Col A(1)=P1, B(2)=Codigo, C(3)=Nombre, D(4)=RFC, E(5)=CURP,
    # F(6)=TipoOperacion, G(7)=CodigoCuenta, H(8)=TasaIVA15...
    # Col Z(26)=TasaIVA11 (índice 25), AA(27)=CodigoCtaGastos (índice 26)

    # Tomar una fila existente como plantilla de valores por defecto
    plantilla = provs_cat.iloc[0].tolist() if len(provs_cat) else []

    for i, prov in enumerate(nuevos_provs):
        nuevo_codigo  = ultimo_codigo + i + 1
        data_row      = [None] * max(30, len(plantilla))

        # Copiar valores de plantilla primero (mantiene tasas y flags estándar)
        for ci, v in enumerate(plantilla):
            if not (isinstance(v, float) and pd.isna(v)):
                data_row[ci] = v

        # Sobreescribir campos clave
        data_row[0]  = "P1"
        data_row[1]  = nuevo_codigo
        data_row[2]  = prov["nombre"]
        data_row[3]  = prov["rfc"]
        data_row[4]  = ""                          # CURP — vacío
        data_row[6]  = prov["cta_proveedor"]       # CodigoCuenta — ya asignada y usada en pólizas
        data_row[26] = CTA_GASTO_DEFAULT            # CodigoCtaGastos — default, verificar

        excel_row = 7 + i
        for ci, v in enumerate(data_row, 1):
            c = ws.cell(excel_row, ci, v)
            c.fill = wfill
            c.font = wfont

        # Nota de verificación en columna después de los datos
        nota_col = max(30, len(plantilla)) + 1
        c = ws.cell(excel_row, nota_col, "⚠ VERIFICAR cuenta de gasto (default 501020000)")
        c.font = nfont

    # Encabezado de nota
    ws.cell(1, max(30, len(plantilla)) + 1, "Nota").font = Font(bold=True, size=9)

    # Ajuste de anchos
    for col in ws.columns:
        max_w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_w + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ── Generación del Catálogo de Cuentas para nuevos proveedores ──

# 9 filas de encabezado estáticas del layout CONTPAq
_HEADERS_CUENTAS = [
    ["Grupo estadístico(E)", "CtaSup"],
    ["Cuenta contable(C)", "Codigo", "Nombre", "NomIdioma", "CtaSup", "Tipo",
     "EsBaja", "CtaMayor", "CtaEfectivo", "FechaRegistro", "SistOrigen",
     "IdMoneda", "DigAgrup", "IdSegNeg", "SegNegMovtos", "Consume", "IdAgrupadorSAT"],
    ["Rubros NIF(RF)", "Codigo"],
    ["Cuenta de flujo de efectivo(F)", "Codigo"],
    ["F", "10200000"],
    ["C", "200000000", "Pasivo", "Liabilities", "000000000", "D", 0, 2, 0, "2019-11-11", 11, "   1", 0, "0", 0, 0, "0"],
    ["C", "200010000", "Pasivo a corto plazo", "Short-term liabilities", "200000000", "D", 0, 3, 0, "2019-11-11", 11, "   1", 0, "0", 0, 0, "0"],
    ["C", "201000000", "Proveedores", "Suppliers", "200010000", "D", 0, 1, 0, "2019-11-11", 11, "   1", 0, "0", 0, 0, "201"],
    ["C", "201010000", "Proveedores nacionales", "Nationals Suppliers", "201000000", "D", 0, 2, 0, "2019-11-11", 11, "   1", 0, "0", 0, 0, "201.01"],
]

# Valores fijos para cada cuenta individual de proveedor
_CUENTA_FIJA = {
    "cta_sup":        "201010000",
    "tipo":           "D",
    "es_baja":        0,
    "cta_mayor":      2,
    "cta_efectivo":   0,
    "sist_origen":    11,
    "id_moneda":      "   1",
    "dig_agrup":      0,
    "id_seg_neg":     "0",
    "seg_neg_movtos": 0,
    "consume":        0,
    "id_agrup_sat":   "201.01",
}


def generar_catalogo_cuentas(nuevos_provs):
    """
    Genera el Excel de catálogo de cuentas CONTPAq para los nuevos proveedores.
    Estructura por cada cuenta:
        C   [numero_cuenta]  [nombre]  201010000  D  0  2  0  [fecha]  11     1  0  0  0  0  201.01
        RF  2102.03
    Las 9 primeras filas son encabezados estáticos fijos.
    """
    from datetime import date, datetime
    fecha_hoy = datetime.combine(date.today(), datetime.min.time())  # datetime real, no string

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalogo Cuentas"

    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", size=9)
    cfill = PatternFill("solid", fgColor="DDEEFF")   # azul claro para cuentas nuevas
    rfill = PatternFill("solid", fgColor="F2F2F2")   # gris para filas RF
    cfont = Font(size=9)

    # ── Filas 1-9: encabezados estáticos ──
    for ri, hrow in enumerate(_HEADERS_CUENTAS, 1):
        for ci, val in enumerate(hrow, 1):
            c = ws.cell(ri, ci, val)
            c.font = hfont
            c.fill = hfill

    # ── Fila 10 en adelante: una cuenta + RF por cada proveedor nuevo ──
    row_idx = 10
    for prov in nuevos_provs:
        # Fila C — cuenta contable
        cuenta_vals = [
            "C",
            str(prov["cta_proveedor"]),          # Codigo
            prov["nombre"],                       # Nombre
            # NomIdioma — vacío (las cuentas individuales no lo tienen)
            None,
            _CUENTA_FIJA["cta_sup"],              # CtaSup = 201010000
            _CUENTA_FIJA["tipo"],                 # D
            _CUENTA_FIJA["es_baja"],              # 0
            _CUENTA_FIJA["cta_mayor"],            # 2
            _CUENTA_FIJA["cta_efectivo"],         # 0
            fecha_hoy,                            # FechaRegistro = hoy
            _CUENTA_FIJA["sist_origen"],          # 11
            _CUENTA_FIJA["id_moneda"],            # "   1"
            _CUENTA_FIJA["dig_agrup"],            # 0
            _CUENTA_FIJA["id_seg_neg"],           # "0"
            _CUENTA_FIJA["seg_neg_movtos"],       # 0
            _CUENTA_FIJA["consume"],              # 0
            _CUENTA_FIJA["id_agrup_sat"],         # "201.01"
        ]
        for ci, val in enumerate(cuenta_vals, 1):
            c = ws.cell(row_idx, ci, val)
            c.fill = cfill
            c.font = cfont
            if ci == 10:   # columna FechaRegistro — formato yyyymmdd como el original
                c.number_format = "yyyymmdd"
        row_idx += 1

        # Fila RF — rubro NIF fijo entre cada cuenta
        ws.cell(row_idx, 1, "RF").fill = rfill
        ws.cell(row_idx, 1).font = cfont
        ws.cell(row_idx, 2, "2102.03").fill = rfill
        ws.cell(row_idx, 2).font = cfont
        row_idx += 1

    # Ajuste de anchos
    anchos = [6, 12, 45, 4, 12, 5, 6, 8, 10, 12, 10, 8, 8, 8, 12, 8, 10]
    for ci, ancho in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
